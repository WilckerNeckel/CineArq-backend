import json
from django.http import JsonResponse
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from requests import Response
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
import io
from django.http import FileResponse
from formulario.models import Formularios
from rest_framework import status


class WorkSheet:
    def __init__(self, ws_file_path=None, codes_file_path=None, worksheet_file=None, codes_file=None):
        self.worksheet_file = worksheet_file
        self.ws_file_path = ws_file_path
        self.main_wb = self.load_worksheet(self.ws_file_path, self.worksheet_file)
        self.main_ws = self.main_wb.active
        self.codes_list = self.return_codes_list(codes_file_path= codes_file_path, codes_file=codes_file)
        self.used_codes = self.create_list_used_codes()
        self.add_column("Inconsistencia")
    
    def save_file(self, new_ws_file_path):
        self.main_wb.save(new_ws_file_path)
    
    def load_worksheet(self, ws_file_path, file):
        self.main_wb = load_workbook(file) if file else load_workbook(ws_file_path)
        return self.main_wb
    
    def load_codes_worksheet(self, ws_file_path, file):
        codes_worksheet = load_workbook(file) if file else load_workbook(ws_file_path)
        return codes_worksheet

    def add_column(self, name_column):
        column_last_letter = self.main_ws.cell(row=1, column=self.main_ws.max_column).column_letter
        new_column_letter = chr(ord(column_last_letter) + 1)
        self.main_ws[f"{new_column_letter}1"] = name_column

    def paint_line(self, number_row):
        red_color = "FFFF00"  # Vermelho
        fill = PatternFill(start_color=red_color, end_color=red_color, fill_type="solid")
        for celula in self.main_ws[number_row]:
            celula.fill = fill
            
    def format_menu(self):
        blue_color = "5da6f4"  # Light blue
        font = Font(name="Arial", bold=True, color=colors.WHITE)
        fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type="solid")
        for cell in self.main_ws[1]:
            cell.font = font
            cell.fill = fill
            
    def format_border(self):
        black_color = "000000"  # Black
        border = Border(left=Side(style="thin", color=black_color),
                        right=Side(style="thin", color=black_color),
                        top=Side(style="thin", color=black_color),
                        bottom=Side(style="thin", color=black_color))
        for row in self.main_ws.iter_rows():
            for cell in row:
                cell.border = border

    def transform_worksheet(self, json_list):
        number_column_code = self.get_number_column_reason(1)
        column = self.get_column_letter(number_column_code)
        
        self.format_menu()
        self.format_border()
        
        for json in json_list:
            for key in json:
                if json[key]["reason"] != "":
                    self.main_ws[f"{column}{key}"].value = json[key]["reason"]
                    self.paint_line(key)
                    
    def get_column_letter(self, column_number):
        return chr(64 + column_number)
    

    def get_number_column_code(self, number_row):
        for celula in self.main_ws[number_row]:
            string_cell = celula.value
            clear_value_cel = string_cell.lower().strip().replace(" ", "_")
            if clear_value_cel == "codigos":
                return celula.column
        
        raise Exception("A coluna 'Codigos' não foi encontrada.")

    def get_number_column_reason(self, number_row):
        for celula in self.main_ws[number_row]:
            string_cell = celula.value
            clear_value_cel = string_cell.lower().strip().replace(" ", "_")
            if clear_value_cel == "inconsistencia":
                return celula.column
        
        raise Exception("A coluna 'Inconsistencia' não foi encontrada.")
    
                
    def create_json_arrays(self):
        json_arrays = []
        column = self.get_number_column_code(1)
        
        for cells in self.main_ws.iter_cols(min_row=1, max_row=self.main_ws.max_row, min_col=column, max_col=column):
            for cell in cells:
                json_arrays.append({cell.row: {"value": cell.value, "row": cell.row, "reason": ""}})
                
                if cell.value not in self.codes_list:
                    #acessa o ultimo elemento da lista
                    json_arrays[-1][cell.row]["reason"] = "Código inválido"
                
                if self.used_codes.count(cell.value) > 1:
                    indices = self.return_index_duplicates(cell.value)
                    json_arrays[-1][cell.row]["reason"] = f"Código duplicado nas linhas {', '.join(str(i) for i in indices)}"

        json_arrays = [json for json in json_arrays if json[list(json.keys())[0]]["value"] != None]
        json_arrays.pop(0)
        return json_arrays
    
    def return_index_duplicates(self, code):
        indices = []
        column = self.get_number_column_code(1)
        for cells in self.main_ws.iter_cols(min_row=1, max_row=self.main_ws.max_row, min_col=column, max_col=column):
            for cell in cells:
                if cell.value == code:
                    indices.append(cell.row)
        return indices
                
    def create_list_used_codes(self):
        used_codes = []
        column = self.get_number_column_code(1)
        
        for cells in self.main_ws.iter_cols(min_row=1, max_row=self.main_ws.max_row, min_col=int(column), max_col=int(column)):
            for cell in cells:
                used_codes.append(cell.value)
        
        used_codes = [code for code in used_codes if code != None]
        used_codes.pop(0)
        return used_codes
                

    def save_file(self, new_ws_file_path):
        self.main_wb.save(new_ws_file_path)
        
    def return_file(self):
        return self.main_wb
        
    def return_codes_list(self, codes_file_path=None, codes_file=None):
        codes_list = []
        
        codes_worksheet = self.load_codes_worksheet(codes_file_path, codes_file)
        codes_worksheet = codes_worksheet.active
        for cell in codes_worksheet['A']:
            codes_list.append(cell.value)
        codes_list.pop(0)
        print(codes_list)
        return codes_list

    def print_codes_list(self):
        print(self.codes_list)
        

class WorkSheetView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:

            worksheet_received = request.FILES['worksheet']
            list_codes_received = request.FILES['list_codes']
            
            
            work_sheet = WorkSheet(worksheet_file=worksheet_received, codes_file=list_codes_received)
            
            json_arrays = work_sheet.create_json_arrays()
            work_sheet.transform_worksheet(json_arrays)
            transformed_worksheet = work_sheet.return_file()

            # Create a file-like object in memory
            output = io.BytesIO()

            # Save the transformed worksheet to the file-like object
            transformed_worksheet.save(output)

            # Set the file-like object's pointer to the beginning
            output.seek(0)

            # Create a file response to return the transformed worksheet
            response = FileResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="planilha_resultante.xlsx"'

            response.status_code = 200
            return response
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        
class gerenerate_worksheet(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            data = json.loads(request.body.decode('utf-8'))
            date = data.get('Data')
            print(date)
            objects = Formularios.objects.filter(created_at__date=date)
            # Create a new workbook
            workbook = Workbook()
            # Get the active sheet
            sheet = workbook.active

            # Write headers
            headers = ['Nome', 'Email', 'CPF', 'Telefone', 'Codigos', 'Data']
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # Write data
            row_num = 2
            for obj in objects:
                data = [obj.nome, obj.email, obj.cpf, obj.telefone, obj.codigo, obj.created_at.strftime("%d/%m/%Y %H:%M:%S")]
                for col_num, value in enumerate(data, 1):
                    sheet.cell(row=row_num, column=col_num, value=value)
                row_num += 1

            output = io.BytesIO()

            workbook.save(output)

            output.seek(0)

            response = FileResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="planilha_resultante.xlsx"'

            response.status_code = 200
            return response

        except Exception as e:
            return JsonResponse(data={'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

    
            